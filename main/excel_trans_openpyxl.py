#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2021/7/8 13:44
# @Author : liu yang
# @Desc: excel 格式转化


from openpyxl import load_workbook


# openpyxl 使用
wb = load_workbook('/Users/liuyang/Downloads/test.xls')
wb.sheetnames
ws = wb['员工刷卡记录表']
ws.dimensions

row_list = []
for row in ws.rows:
    for cell in row:
        # print(cell.value)
        if cell.value is not None:
            if "工号" in str(cell.value):
                # print(ws.cell(row=cell.row, column=2).value)
                row_list.append(cell.row)
                # print(cell.row, cell.column)
                # print(row_list)
                break
print(row_list)

